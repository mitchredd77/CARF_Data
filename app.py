import openpyxl, shutil, os, time
from openpyxl import load_workbook
import pandas as pd

wb1 = load_workbook(filename="PatientCARF_Data.xlsx")
sheet = wb1.active

# Removes files from previous runs if they exist, copies blank template of Patient Data before running
if os.path.exists("New_Patient_Data.xlsx"):
   print("New_Patient_Data.xlsx exists")
   os.remove("New_Patient_Data.xlsx")
   shutil.copy("Template/New_Patient_Data.xlsx", "New_Patient_Data.xlsx")
else: 
    shutil.copy("Template/New_Patient_Data.xlsx", "New_Patient_Data.xlsx")
if os.path.exists("Template/New_Patient_Data_Clean.xlsx"):
    os.remove("Template/New_Patient_Data_Clean.xlsx")

wb2 = load_workbook(filename="New_Patient_Data.xlsx")
sheet2 = wb2.active

# Finds the correct column by searching headers
def find_col(header):
    column_index = 0
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        for cell in col:
            if cell.value == header:
                column_index = cell.column
                break
    return column_index
#
def get_all_dx_codes():
    total_dx = []
    col = find_col("Dx Codes (From Claim)")
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is None:
            continue
        else:
            dx_codes = sheet.cell(row=row, column=col).value
            dx_codes = dx_codes.split(", ")
            for code in dx_codes:
                if code in total_dx:
                    continue
                else:
                    total_dx.append(code)
    return total_dx

def menu_which_dx_code():
    menu = True
    while menu == True:
         print("""
               ****************************
               Welcome to the Dx CODE FINDER!
               ****************************""")
         choice = input("""
                        *************************************
                           Which Dx code are you looking for? 
                        *************************************
                        ___""")
         total_dx_codes = get_all_dx_codes()
         if choice in total_dx_codes:
             return choice
             menu = False
         else:
             print(", ".join(total_dx_codes))
             input("""
                   **Please choose a Dx code from above:**
                   """)
    return choice

def add_data_dx_codes(dx_code):
    col = find_col("Dx Codes (From Claim)")
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value == None:
            continue
        else:
            dx_codes = sheet.cell(row=row, column=col).value
            dx_codes = dx_codes.split(", ")
            if dx_code in dx_codes:
                for i in range(1, 11):
                    sheet2.cell(row=row, column=i).value = sheet.cell(row=row, column=i).value
                
def clean_sheet(input_file, output_file):
    # Read the spreadsheet into a DataFrame
    df = pd.read_excel(input_file)  # Change this line if your spreadsheet is in a different format

    # Remove rows with all NaN (empty) values
    df_cleaned = df.dropna(how='all')

    # Save the cleaned DataFrame to a new spreadsheet
    df_cleaned.to_excel(output_file, index=False)  # Change this line if you want to save in a different format
    


def app():
    choice = menu_which_dx_code()
    add_data_dx_codes(choice)
    wb2.save("New_Patient_Data.xlsx")
    clean_sheet("New_Patient_Data.xlsx", "New_Patient_Data_Clean.xlsx")
    print("""
                  ****************************************************
                  ALL DATA HAS BEEN WRITTEN TO THE New_Patient_Data.xlsx file
                                HAVE A WONDERFUL DAY
                  ****************************************************""")
    time.sleep(1.5)
if __name__ == '__main__':
      app()
    
            


        